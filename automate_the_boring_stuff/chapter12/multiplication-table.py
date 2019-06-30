#! python3
'''
    Takes a number 'n' from the command line and creates an excel
    multiplication table
'''
import sys, openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font

def multiplicationTable(n):
    # Create workbook
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = str(n) + ' Multiplication Table'

    # Initialize numbers ranging from 1 to n in 1st col and row
    for num in range(1, n + 1):
        sheet.cell(row=1, column=num+1).value = num
        sheet.cell(row=1, column=num+1).font = Font(bold=True)

        sheet.cell(row=num+1, column=1).value = num
        sheet.cell(row=num+1, column=1).font = Font(bold=True)

    # TODO: Compute the numbers in the table

    for row in range(1, n + 1):
        for col in range(1, n + 1):
            sheet.cell(row=row, column=col).value = row * col


    wb.save(f'{str(n)} Multiplication Table.xlsx')

if len(sys.argv) < 2:
    print('enter value for n')
else:
    n = sys.argv[1]
    multiplicationTable(int(n))